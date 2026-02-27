"""엑셀/PDF 내보내기 모듈"""

import io
from datetime import date, timedelta
from pathlib import Path

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# 한글 폰트 경로 (backend/assets/)
_ASSETS_DIR = Path(__file__).resolve().parent.parent / "assets"
_FONT_REGULAR = str(_ASSETS_DIR / "malgun.ttf")
_FONT_BOLD = str(_ASSETS_DIR / "malgunbd.ttf")


def export_analyses_to_excel(queryset) -> io.BytesIO:
    """분석 결과를 .xlsx 파일로 내보내기"""
    wb = Workbook()
    ws = wb.active
    ws.title = "분석 결과"

    # 헤더 스타일
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [
        "No",
        "케이스 ID",
        "기사 제목",
        "언론사",
        "게재일",
        "적합도",
        "판단 근거",
        "사건 분야",
        "상대방",
        "피해 규모",
        "피해자 수",
        "진행 단계",
        "진행 상세",
        "요약",
        "원문 링크",
    ]

    # 헤더 작성
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # 적합도 배경색
    suit_fills = {
        "High": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
        "Medium": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
        "Low": PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid"),
    }

    # 데이터 작성
    for idx, analysis in enumerate(queryset, 1):
        article = analysis.article
        case_id = analysis.case_group.case_id if analysis.case_group else ""
        row = idx + 1

        values = [
            idx,
            case_id,
            article.title,
            article.source.name if article.source else "",
            article.published_at.strftime("%Y-%m-%d") if article.published_at else "",
            analysis.suitability,
            analysis.suitability_reason,
            analysis.case_category,
            analysis.defendant,
            analysis.damage_amount,
            analysis.victim_count,
            analysis.stage,
            analysis.stage_detail,
            analysis.summary,
            article.url,
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        # 적합도 셀 색상
        suit_cell = ws.cell(row=row, column=6)
        fill = suit_fills.get(analysis.suitability)
        if fill:
            suit_cell.fill = fill

    # 열 너비 조정
    col_widths = [5, 16, 40, 12, 12, 8, 50, 15, 15, 15, 12, 12, 20, 60, 40]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # 첫 행 고정
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_case_groups_to_excel(queryset) -> io.BytesIO:
    """사건 그룹(CaseGroup)을 .xlsx 파일로 내보내기 (사건 단위)"""
    wb = Workbook()
    ws = wb.active
    ws.title = "사건 목록"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [
        "No",
        "케이스 ID",
        "사건명",
        "기사 수",
        "심사 결과",
        "심사 완료",
        "수임 통과",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    suit_fills = {
        "High": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
        "Medium": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
        "Low": PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid"),
    }

    for idx, cg in enumerate(queryset, 1):
        row = idx + 1
        article_count = getattr(cg, "article_count", None) or (cg.analyses.filter(is_relevant=True).count() if hasattr(cg, "analyses") else 0)
        values = [
            idx,
            cg.case_id or "",
            cg.name or "",
            article_count,
            cg.client_suitability or "",
            "✓" if cg.review_completed else "—",
            "✓" if cg.accepted else "—",
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if cg.client_suitability:
            suit_cell = ws.cell(row=row, column=5)
            fill = suit_fills.get(cg.client_suitability)
            if fill:
                suit_cell.fill = fill

    col_widths = [5, 16, 45, 10, 12, 12, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ──────────────────────────────────────────────
# PDF 내보내기
# ──────────────────────────────────────────────

# 적합도 → 색상 (R, G, B)
_SUIT_COLORS = {
    "High": (220, 38, 38),    # 빨강
    "Medium": (234, 179, 8),  # 노랑
    "Low": (107, 114, 128),   # 회색
}


class _KoreanPDF(FPDF):
    """맑은 고딕 폰트가 등록된 PDF 기반 클래스"""

    def __init__(self, period_label: str, generated_at: str):
        super().__init__(orientation="L", unit="mm", format="A4")
        self.period_label = period_label
        self.generated_at = generated_at
        self.add_font("Malgun", style="", fname=_FONT_REGULAR)
        self.add_font("Malgun", style="B", fname=_FONT_BOLD)
        self.set_auto_page_break(auto=True, margin=15)

    def header(self):
        self.set_font("Malgun", "B", 9)
        self.set_text_color(100, 100, 100)
        self.cell(0, 6, f"LawNGood 분석 리포트  |  {self.period_label}", align="L")
        self.ln(8)
        # 구분선
        self.set_draw_color(200, 200, 200)
        self.line(10, self.get_y(), self.w - 10, self.get_y())
        self.ln(3)

    def footer(self):
        self.set_y(-12)
        self.set_font("Malgun", "", 8)
        self.set_text_color(150, 150, 150)
        self.cell(0, 5, f"생성일시: {self.generated_at}  |  페이지 {self.page_no()}", align="C")


def _period_range(period: str) -> tuple[date, date, str]:
    """period='weekly'|'monthly' → (date_from, date_to, label)"""
    today = date.today()
    if period == "monthly":
        date_from = today.replace(day=1)
        date_to = today
        label = today.strftime("%Y년 %m월")
    else:  # weekly (기본)
        date_from = today - timedelta(days=6)
        date_to = today
        label = f"{date_from.strftime('%Y.%m.%d')} ~ {date_to.strftime('%m.%d')} (주간)"
    return date_from, date_to, label


def export_analyses_to_pdf(queryset, period: str = "weekly") -> io.BytesIO:
    """분석 결과를 PDF 파일로 내보내기

    Args:
        queryset: Analysis QuerySet (이미 필터링된 상태)
        period: 'weekly' 또는 'monthly'
    Returns:
        BytesIO PDF 바이너리
    """
    _, _, label = _period_range(period)
    generated_at = date.today().strftime("%Y-%m-%d")

    pdf = _KoreanPDF(period_label=label, generated_at=generated_at)

    # ── 표지 ──────────────────────────────────────
    pdf.add_page()
    pdf.set_font("Malgun", "B", 28)
    pdf.set_text_color(15, 23, 42)  # slate-900
    pdf.ln(20)
    pdf.cell(0, 14, "LawNGood 분석 리포트", align="C")
    pdf.ln(12)
    pdf.set_font("Malgun", "", 16)
    pdf.set_text_color(71, 85, 105)  # slate-500
    pdf.cell(0, 10, label, align="C")
    pdf.ln(6)
    pdf.set_font("Malgun", "", 11)
    pdf.cell(0, 8, f"생성일: {generated_at}", align="C")

    # 요약 통계
    analyses = list(queryset)
    total = len(analyses)
    high_cnt = sum(1 for a in analyses if a.suitability == "High")
    medium_cnt = sum(1 for a in analyses if a.suitability == "Medium")
    accepted_cnt = sum(1 for a in analyses if a.accepted)

    pdf.ln(20)
    pdf.set_draw_color(226, 232, 240)

    # 통계 박스 4개 (가로 배열) — start_y 고정으로 계단 방지
    box_w = 55
    box_h = 22
    start_x = (pdf.w - box_w * 4 - 6 * 3) / 2
    start_y = pdf.get_y()
    stats = [
        ("대상 기사", str(total), (59, 130, 246)),
        ("High 적합", str(high_cnt), (220, 38, 38)),
        ("Medium 적합", str(medium_cnt), (234, 179, 8)),
        ("심사 통과", str(accepted_cnt), (34, 197, 94)),
    ]
    for i, (title, value, color) in enumerate(stats):
        x = start_x + i * (box_w + 6)
        y = start_y
        pdf.set_fill_color(248, 250, 252)
        pdf.rect(x, y, box_w, box_h, style="F")
        pdf.set_font("Malgun", "B", 18)
        pdf.set_text_color(*color)
        pdf.set_xy(x, y + 2)
        pdf.cell(box_w, 10, value, align="C")
        pdf.set_font("Malgun", "", 9)
        pdf.set_text_color(100, 116, 139)
        pdf.set_xy(x, y + 13)
        pdf.cell(box_w, 6, title, align="C")

    # ── 사건 목록 테이블 ───────────────────────────
    pdf.add_page()
    pdf.set_font("Malgun", "B", 11)
    pdf.set_text_color(15, 23, 42)
    pdf.cell(0, 8, "사건 목록")
    pdf.ln(10)

    # 열 너비 (가로 A4 = 297mm, 여백 각 10mm)
    col_widths = [8, 22, 68, 18, 20, 16, 28, 28, 58]
    col_headers = ["No", "케이스 ID", "기사 제목", "언론사", "게재일", "적합도", "사건 분야", "피해 규모", "요약"]
    row_h = 8

    # 테이블 헤더
    pdf.set_fill_color(15, 23, 42)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Malgun", "B", 8)
    for w, h in zip(col_widths, col_headers):
        pdf.cell(w, row_h, h, border=1, fill=True, align="C")
    pdf.ln()

    # 데이터 행
    pdf.set_font("Malgun", "", 7)
    for idx, analysis in enumerate(analyses, 1):
        article = analysis.article
        suit = analysis.suitability or ""
        r, g, b = _SUIT_COLORS.get(suit, (107, 114, 128))
        case_id = analysis.case_group.case_id if analysis.case_group else "-"

        # 배경색 교대
        if idx % 2 == 0:
            pdf.set_fill_color(248, 250, 252)
        else:
            pdf.set_fill_color(255, 255, 255)

        fill = True
        pdf.set_text_color(30, 30, 30)

        row_data = [
            (str(idx), "C"),
            (case_id, "C"),
            (article.title[:38] + "…" if len(article.title) > 38 else article.title, "L"),
            (article.source.name[:8] if article.source else "-", "C"),
            (article.published_at.strftime("%y.%m.%d") if article.published_at else "-", "C"),
            (suit, "C"),
            ((analysis.case_category or "-")[:12], "C"),
            ((analysis.damage_amount or "-")[:12], "C"),
            ((analysis.summary or "")[:42] + "…" if len(analysis.summary or "") > 42 else (analysis.summary or "-"), "L"),
        ]

        for i, ((text, align), w) in enumerate(zip(row_data, col_widths)):
            # 적합도 셀은 색상 강조
            if i == 5:
                pdf.set_text_color(r, g, b)
                pdf.set_font("Malgun", "B", 7)
                pdf.cell(w, row_h, text, border=1, align=align, fill=fill)
                pdf.set_text_color(30, 30, 30)
                pdf.set_font("Malgun", "", 7)
            else:
                pdf.cell(w, row_h, text, border=1, align=align, fill=fill)
        pdf.ln()

    return io.BytesIO(pdf.output())
